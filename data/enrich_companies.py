#!/usr/bin/env python3
"""
Company Contact Enricher (Hybrid: HTML + Browser + Google Places + Facebook)
===========================================================================
Reads company names from a spreadsheet, finds their website, email,
phone number, and LinkedIn profile, and writes results to a new spreadsheet.

Hybrid approach with fallbacks:
  1. HTML scraping (fast)
  2. Browser automation (JS-heavy sites)
  3. Google Places API (official business data)
  4. Facebook business pages (last resort)

Usage:
    export GOOGLE_PLACES_API_KEY="your_key_here"
    python enrich_companies.py input.xlsx
    python enrich_companies.py "https://docs.google.com/spreadsheets/d/..." --column "Company Name"

To get Google Places API key:
  1. Go to https://console.cloud.google.com/
  2. Create a new project
  3. Enable "Places API"
  4. Create API credentials (API key)
  5. Set environment variable: GOOGLE_PLACES_API_KEY="your_key"

Requirements:
    pip install requests beautifulsoup4 openpyxl pandas tqdm playwright googlemaps facebook-scraper
"""

import argparse
import re
import time
import random
import logging
from pathlib import Path
from urllib.parse import urljoin, urlparse, unquote, parse_qs
from concurrent.futures import ThreadPoolExecutor, as_completed

import pandas as pd
import requests
from bs4 import BeautifulSoup
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment
from tqdm import tqdm

try:
    from playwright.sync_api import sync_playwright
    PLAYWRIGHT_AVAILABLE = True
except ImportError:
    PLAYWRIGHT_AVAILABLE = False

try:
    import googlemaps
    GOOGLE_PLACES_AVAILABLE = True
except ImportError:
    GOOGLE_PLACES_AVAILABLE = False

try:
    from facebook_scraper import get_page  # type: ignore[import]
    FACEBOOK_AVAILABLE = True
except ImportError:
    FACEBOOK_AVAILABLE = False

# ─── Configuration ─────────────────────────────────────────────────────────────

DELAY_BETWEEN_REQUESTS = (1.5, 3.5)   # seconds (min, max) — be polite
REQUEST_TIMEOUT        = 10            # seconds per HTTP request
MAX_RETRIES            = 2
MAX_WORKERS            = 4             # parallel workers for enrichment
GOOGLE_PLACES_API_KEY  = "AIzaSyCl7nu1P05fRtWwe3HWXAjngPcpwhRVWl0"
USER_AGENT = (
    "Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
    "AppleWebKit/537.36 (KHTML, like Gecko) "
    "Chrome/124.0.0.0 Safari/537.36"
)

logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s  %(levelname)-7s  %(message)s",
    datefmt="%H:%M:%S",
)
log = logging.getLogger(__name__)

# ── HTTP helpers ───────────────────────────────────────────────────────────────

SESSION = requests.Session()
SESSION.headers.update({"User-Agent": USER_AGENT})


def _get(url: str, **kwargs) -> requests.Response | None:
    for attempt in range(MAX_RETRIES + 1):
        try:
            r = SESSION.get(url, timeout=REQUEST_TIMEOUT, **kwargs)
            r.raise_for_status()
            return r
        except requests.RequestException as e:
            if attempt < MAX_RETRIES:
                time.sleep(2 ** attempt)
            else:
                log.debug(f"GET failed for {url}: {e}")
    return None


def _sleep():
    time.sleep(random.uniform(*DELAY_BETWEEN_REQUESTS))


# ── Web search (DuckDuckGo HTML — no API key needed) ──────────────────────────

def search_web(query: str, num_results: int = 5) -> list[str]:
    """Return a list of result URLs from DuckDuckGo for the given query."""
    url = "https://html.duckduckgo.com/html/"
    try:
        r = SESSION.post(
            url,
            data={"q": query, "b": "", "kl": ""},
            timeout=REQUEST_TIMEOUT,
            headers={"User-Agent": USER_AGENT},
        )
        soup = BeautifulSoup(r.text, "html.parser")
        links = []
        for a in soup.select("a.result__a"):
            href = a.get("href", "")
            if not href:
                continue
            # DuckDuckGo wraps links as //duckduckgo.com/l/?uddg=ENCODED_URL
            if "uddg=" in href:
                params = parse_qs(urlparse(href).query)
                if "uddg" in params:
                    href = unquote(params["uddg"][0])
            if href.startswith("http"):
                links.append(href)
            if len(links) >= num_results:
                break
        return links
    except Exception as e:
        log.debug(f"DuckDuckGo search failed: {e}")
        return []


def search_emails_from_snippets(query: str) -> list[str]:
    """Search DuckDuckGo and extract any email addresses found in result snippets."""
    url = "https://html.duckduckgo.com/html/"
    try:
        r = SESSION.post(
            url,
            data={"q": query, "b": "", "kl": ""},
            timeout=REQUEST_TIMEOUT,
            headers={"User-Agent": USER_AGENT},
        )
        soup = BeautifulSoup(r.text, "html.parser")
        # Scan full result block text (snippets + titles) for emails
        results_text = soup.get_text(" ", strip=True)
        emails = [e for e in EMAIL_RE.findall(results_text) if _is_valid_email(e)]
        return list(dict.fromkeys(emails))  # deduplicate preserving order
    except Exception as e:
        log.debug(f"DuckDuckGo snippet email search failed: {e}")
        return []


# ── Website discovery ──────────────────────────────────────────────────────────

SKIP_DOMAINS = {
    "linkedin.com", "facebook.com", "twitter.com", "instagram.com",
    "youtube.com", "wikipedia.org", "yelp.com", "crunchbase.com",
    "bloomberg.com", "reuters.com", "glassdoor.com", "indeed.com",
    "zoominfo.com", "dnb.com", "trustpilot.com",
}

CONTACT_PAGE_KEYWORDS = re.compile(
    r"contact|about|reach|touch|enquir|inquir|support|help|team|office|location|connect|address|headquarter|hq|customer|service",
    re.IGNORECASE,
)
MAX_CONTACT_PAGES = 12


def _normalize_url(href: str, base: str) -> str:
    return urljoin(base, href.split("#", 1)[0].strip())


def _is_same_domain(url: str, base: str) -> bool:
    parsed_url = urlparse(url)
    parsed_base = urlparse(base)
    return parsed_url.netloc == parsed_base.netloc


_LEGAL_SUFFIXES = re.compile(
    r"\b(limited|ltd\.?|llp|plc|inc\.?|llc|gmbh|co\.?|contracts?|services?|solutions?|group|holdings?)\b",
    re.IGNORECASE,
)


def _company_keywords(company_name: str) -> list[str]:
    """Return meaningful words from a company name for ownership verification."""
    stripped = _LEGAL_SUFFIXES.sub("", company_name)
    words = re.findall(r"[a-zA-Z0-9]{3,}", stripped)
    return [w.lower() for w in words]


def verify_website_ownership(url: str, company_name: str) -> bool:
    """Return True if the company name's key words appear on the webpage."""
    keywords = _company_keywords(company_name)
    if not keywords:
        return True  # Can't verify — allow through

    r = _get(url)
    if not r:
        return False

    soup = BeautifulSoup(r.text, "html.parser")
    for tag in soup(["script", "style"]):
        tag.decompose()
    page_text = soup.get_text(" ", strip=True).lower()

    # Also check the domain itself — e.g. "ultratherm.com" confirms "ultratherm"
    domain = urlparse(url).netloc.lower().replace("www.", "")

    matches = sum(1 for kw in keywords if kw in page_text or kw in domain)
    ratio = matches / len(keywords)
    passed = ratio >= 0.5 or matches >= 2

    log.debug(
        f"Ownership check '{company_name}' → {url}: "
        f"{matches}/{len(keywords)} keywords matched (ratio {ratio:.2f}) → {'PASS' if passed else 'FAIL'}"
    )
    return passed


def _guess_domains(company_name: str) -> list[str]:
    """Generate a small set of likely domain names to try directly (max 5)."""
    keywords = _company_keywords(company_name)
    if not keywords:
        return []
    slug  = "".join(keywords[:2])   # e.g. "ultratherm"
    slug1 = keywords[0]             # first keyword only
    # Limit to 5 attempts: most NI businesses use .co.uk or .com
    return [
        f"https://www.{slug}.co.uk",
        f"https://www.{slug}.com",
        f"https://www.{slug1}.co.uk",
        f"https://www.{slug1}.com",
        f"https://www.{slug1}.ie",
    ]


def find_website(company_name: str, location: str = "") -> str:
    """Search for the company's official website, verified to actually belong to the company."""
    loc = f" {location}" if location else ""
    # Use flexible queries — full legal name in quotes is often too strict
    core = " ".join(_company_keywords(company_name))
    queries = [
        f'"{company_name}" official website',
        f"{core}{loc} company website",
        f"{core}{loc} contact",
    ]
    for query in queries:
        results = search_web(query, num_results=6)
        for url in results:
            domain = urlparse(url).netloc.lower().replace("www.", "")
            if any(skip in domain for skip in SKIP_DOMAINS):
                continue
            if verify_website_ownership(url, company_name):
                return url
        _sleep()

    # Fallback: try guessed domains directly when search returns nothing
    for url in _guess_domains(company_name):
        r = _get(url)
        if r and verify_website_ownership(r.url, company_name):
            log.debug(f"Domain guess succeeded for {company_name}: {r.url}")
            return r.url

    return ""


# ── Contact info extraction from website ──────────────────────────────────────

EMAIL_RE   = re.compile(r"[a-zA-Z0-9._%+\-]+@[a-zA-Z0-9.\-]+\.[a-zA-Z]{2,}")
PHONE_RE   = re.compile(
    r"(?:(?:\+|00)44[\s\-]?)?(?:0[\s\-]?)?(?:\d[\s\-]?){9,11}"  # UK
    r"|(?:\+?1[\s\-.]?)?\(?\d{3}\)?[\s\-.]?\d{3}[\s\-.]?\d{4}"  # US
    r"|(?:\+\d{1,3}[\s\-.]?)?\(?\d{2,4}\)?[\s\-.]?\d{3,4}[\s\-.]?\d{3,4}"  # International
)

SPAM_EMAILS = {
    "example.com", "domain.com", "email.com", "yourmail.com",
    "sentry.io", "wixpress.com", "squarespace.com", "wordpress.com",
    "googletagmanager.com", "schema.org", "w3.org",
}

SPAM_EMAIL_PREFIXES = {
    "noreply", "no-reply", "donotreply", "do-not-reply",
    "webmaster", "postmaster", "mailer-daemon",
}


def _clean_phone(raw: str) -> str:
    digits = re.sub(r"[^\d+]", "", raw)
    return raw.strip() if len(digits) >= 7 else ""


def _is_valid_email(email: str) -> bool:
    parts = email.lower().split("@")
    if len(parts) != 2:
        return False
    local, domain = parts
    if domain in SPAM_EMAILS:
        return False
    if local in SPAM_EMAIL_PREFIXES:
        return False
    if email.endswith((".png", ".jpg", ".gif", ".svg", ".css", ".js")):
        return False
    # Must have a real TLD
    if "." not in domain:
        return False
    return True


def extract_contacts_from_page(html: str, base_domain: str) -> dict:
    soup = BeautifulSoup(html, "html.parser")

    emails: list[str] = []
    phones: list[str] = []

    # ── Priority 1: mailto: and tel: href links (most reliable) ──────────────
    for a in soup.find_all("a", href=True):
        href = a["href"].strip()
        if href.startswith("mailto:"):
            email = href[7:].split("?")[0].strip()
            if email and _is_valid_email(email):
                emails.append(email)
        elif href.startswith("tel:"):
            raw = href[4:].strip().replace("%20", " ")
            cleaned = _clean_phone(raw)
            if cleaned:
                phones.append(cleaned)

    # ── Priority 2: Schema.org / JSON-LD structured data ─────────────────────
    for script in soup.find_all("script", type="application/ld+json"):
        try:
            import json
            data = json.loads(script.string or "")
            # Flatten nested structures
            def _extract_ld(obj):
                if isinstance(obj, dict):
                    for k, v in obj.items():
                        if k in ("email", "contactEmail") and isinstance(v, str) and _is_valid_email(v):
                            emails.append(v)
                        elif k in ("telephone", "phone", "faxNumber") and isinstance(v, str):
                            c = _clean_phone(v)
                            if c:
                                phones.append(c)
                        else:
                            _extract_ld(v)
                elif isinstance(obj, list):
                    for item in obj:
                        _extract_ld(item)
            _extract_ld(data)
        except Exception:
            pass

    # ── Priority 3: Plain text scan (fallback) ────────────────────────────────
    # Remove script/style noise before scanning
    for tag in soup(["script", "style", "noscript"]):
        tag.decompose()
    text = soup.get_text(" ", strip=True)

    for e in EMAIL_RE.findall(text):
        if _is_valid_email(e) and e not in emails:
            emails.append(e)

    for raw_phone in PHONE_RE.findall(text):
        cleaned = _clean_phone(raw_phone)
        if cleaned and cleaned not in phones:
            phones.append(cleaned)

    # ── Pick best email: prefer domain-matching, then any valid ───────────────
    domain_emails = [e for e in emails if base_domain in e.lower()]
    best_email = (domain_emails or emails or [""])[0]

    # ── Pick best phone: tel: href links first (explicit), then text scan ─────
    # phones list is ordered: Priority 1 (tel: links) → Priority 2 (JSON-LD) → Priority 3 (text)
    # Prefer the first tel:-sourced phone; only fall back to text-scan if none found via href/LD
    best_phone = phones[0] if phones else ""

    return {"email": best_email, "phone": best_phone}


def _discover_contact_pages(soup: BeautifulSoup, base: str, known_urls: set[str] | None = None) -> list[str]:
    """Find links to contact/about pages within the site."""
    known_urls = set(known_urls or [])
    found = []

    for a in soup.find_all("a", href=True):
        href = a["href"].strip()
        if href.startswith(("mailto:", "tel:", "#", "javascript:")):
            continue

        full = _normalize_url(href, base)
        if not full or full in known_urls:
            continue
        if not _is_same_domain(full, base):
            continue

        text = a.get_text(" ", strip=True).lower()
        if CONTACT_PAGE_KEYWORDS.search(href) or CONTACT_PAGE_KEYWORDS.search(text):
            found.append(full)
            known_urls.add(full)

    return found


def scrape_website(website_url: str) -> dict:
    """Scrape homepage and contact-related pages for email and phone."""
    result = {"email": "", "phone": ""}
    if not website_url:
        return result

    parsed   = urlparse(website_url)
    base     = f"{parsed.scheme}://{parsed.netloc}"
    base_dom = parsed.netloc.lower().replace("www.", "")

    # Canonical contact page paths to always try
    contact_paths = [
        "/contact",
        "/contact-us",
        "/contact_us",
        "/contacts",
        "/contactus",
        "/about",
        "/about-us",
        "/about_us",
        "/about/contact",
        "/reach-us",
        "/get-in-touch",
        "/getintouch",
        "/enquire",
        "/enquiries",
        "/support",
        "/help",
        "/team",
        "/our-team",
        "/location",
        "/locations",
        "/offices",
    ]
    pages_to_try = [website_url] + [urljoin(base, p) for p in contact_paths]
    known_urls = set(pages_to_try)
    visited = set()

    for page_url in pages_to_try:
        if page_url in visited:
            continue
        visited.add(page_url)

        r = _get(page_url)
        if not r:
            continue

        soup = BeautifulSoup(r.text, "html.parser")

        if page_url == website_url or len(pages_to_try) < MAX_CONTACT_PAGES:
            discovered = _discover_contact_pages(soup, base, known_urls)
            for d in discovered:
                if len(pages_to_try) >= MAX_CONTACT_PAGES:
                    break
                if d not in known_urls:
                    known_urls.add(d)
                    pages_to_try.append(d)

        found = extract_contacts_from_page(r.text, base_dom)
        if found["email"] and not result["email"]:
            result["email"] = found["email"]
        if found["phone"] and not result["phone"]:
            result["phone"] = found["phone"]
        if result["email"] and result["phone"]:
            break
        _sleep()

    return result


def scrape_website_browser(website_url: str) -> dict:
    """Scrape website using browser automation (fallback for JS-heavy sites)."""
    if not PLAYWRIGHT_AVAILABLE or not website_url:
        return {"email": "", "phone": ""}
    
    result = {"email": "", "phone": ""}
    try:
        with sync_playwright() as p:
            browser = p.chromium.launch(headless=True, timeout=10000)
            context = browser.new_context(
                user_agent=USER_AGENT,
                viewport={"width": 1280, "height": 720},
                ignore_https_errors=True
            )
            page = context.new_page()
            
            parsed = urlparse(website_url)
            base = f"{parsed.scheme}://{parsed.netloc}"
            base_dom = parsed.netloc.lower().replace("www.", "")

            contact_paths = [
                "/contact", "/contact-us", "/contact_us", "/contacts",
                "/about", "/about-us", "/reach-us", "/get-in-touch",
                "/enquire", "/enquiries", "/support", "/team",
            ]
            pages_to_try = [urljoin(base, p) for p in contact_paths] + [website_url]
            visited = set()
            
            for page_url in pages_to_try:
                if page_url in visited:
                    continue
                visited.add(page_url)
                try:
                    page.goto(page_url, wait_until="domcontentloaded", timeout=8000)
                    time.sleep(0.5)  # Let JS render
                    html = page.content()
                    soup = BeautifulSoup(html, "html.parser")

                    if len(pages_to_try) < MAX_CONTACT_PAGES:
                        discovered = _discover_contact_pages(soup, base, set(pages_to_try))
                        for d in discovered:
                            if len(pages_to_try) >= MAX_CONTACT_PAGES:
                                break
                            if d not in pages_to_try:
                                pages_to_try.append(d)

                    found = extract_contacts_from_page(html, base_dom)
                    if found["email"] and not result["email"]:
                        result["email"] = found["email"]
                    if found["phone"] and not result["phone"]:
                        result["phone"] = found["phone"]
                    if result["email"] and result["phone"]:
                        break
                except Exception as e:
                    log.debug(f"Browser scrape failed for {page_url}: {e}")
                    continue
            
            browser.close()
    except Exception as e:
        log.debug(f"Browser automation failed for {website_url}: {e}")
    
    return result



# ── Google Places API lookup ────────────────────────────────────────────────────

def _name_similarity(a: str, b: str) -> float:
    """Return 0–1 similarity score between two company name strings."""
    from difflib import SequenceMatcher
    # Strip legal suffixes before comparing
    _strip = re.compile(r"\b(limited|ltd\.?|llp|plc|inc\.?|llc|gmbh|co\.?)\b", re.IGNORECASE)
    a = _strip.sub("", a).strip().lower()
    b = _strip.sub("", b).strip().lower()
    return SequenceMatcher(None, a, b).ratio()


def lookup_google_places(company_name: str, location: str = "UK", threshold: float = 0.45) -> dict:
    """Look up company on Google Places API to find phone and other info."""
    if not GOOGLE_PLACES_AVAILABLE or not GOOGLE_PLACES_API_KEY:
        return {"phone": "", "email": "", "website": ""}

    try:
        gmaps = googlemaps.Client(key=GOOGLE_PLACES_API_KEY, timeout=10)
        query = f"{company_name} {location}"

        results = gmaps.places(query, type="establishment")

        # Find best-matching result by name similarity — never blindly take top result
        best_place = None
        best_score = 0.0
        for candidate in results.get("results", [])[:5]:
            score = _name_similarity(company_name, candidate.get("name", ""))
            if score > best_score:
                best_score = score
                best_place = candidate

        # Require minimum name similarity to avoid returning a wrong company
        if not best_place or best_score < threshold:
            log.debug(f"Google Places: no confident match for '{company_name}' (best score {best_score:.2f})")
            return {"phone": "", "email": "", "website": ""}

        place = best_place
        log.debug(f"Google Places matched '{place.get('name')}' for '{company_name}' (score {best_score:.2f})")
        phone = place.get("formatted_phone_number", "")
        website = place.get("website", "")

        place_id = place.get("place_id")
        if place_id:
            details = gmaps.place(place_id, fields=[
                "formatted_phone_number", "website", "formatted_address"
            ])
            if details.get("result"):
                phone = details["result"].get("formatted_phone_number", phone)
                website = details["result"].get("website", website)

        return {"phone": phone, "email": "", "website": website}
    except Exception as e:
        log.debug(f"Google Places lookup failed for {company_name}: {e}")

    return {"phone": "", "email": "", "website": ""}


# ── Facebook business page scraper ────────────────────────────────────────────

def scrape_facebook_business(company_name: str, timeout: int = 20) -> dict:
    """Scrape Facebook business page for contact info (hard timeout via thread)."""
    if not FACEBOOK_AVAILABLE:
        return {"phone": "", "email": ""}
    from concurrent.futures import ThreadPoolExecutor, TimeoutError as FutureTimeout
    with ThreadPoolExecutor(max_workers=1) as _ex:
        _fut = _ex.submit(_scrape_facebook_business_impl, company_name)
        try:
            return _fut.result(timeout=timeout)
        except (FutureTimeout, Exception) as e:
            log.debug(f"Facebook scrape timed out or failed for {company_name}: {e}")
            return {"phone": "", "email": ""}


def _scrape_facebook_business_impl(company_name: str) -> dict:
    
    try:
        # Search for Facebook page URL first
        search_query = f"{company_name} site:facebook.com"
        results = search_web(search_query, num_results=3)
        
        for fb_url in results:
            if "facebook.com" in fb_url and "/pages/" in fb_url:
                try:
                    # Extract page name from URL
                    page_match = re.search(r'/pages/([^/?]+)', fb_url)
                    if page_match:
                        page_name = page_match.group(1)
                        page = get_page(page_name, timeout=10)
                        
                        phone = page.get("phone", "")
                        email = page.get("email", "")
                        
                        if phone or email:
                            log.debug(f"Facebook found for {company_name}: {phone or email}")
                            return {"phone": phone, "email": email}
                except Exception as e:
                    log.debug(f"Facebook scrape failed for {fb_url}: {e}")
                    continue
    except Exception as e:
        log.debug(f"Facebook lookup failed for {company_name}: {e}")
    
    return {"phone": "", "email": ""}


# ── LinkedIn search ────────────────────────────────────────────────────────────

def find_linkedin(company_name: str) -> str:
    """Return a LinkedIn company page URL if found."""
    query = f"site:linkedin.com/company {company_name}"
    results = search_web(query, num_results=5)
    for url in results:
        if "linkedin.com/company/" in url:
            # Normalise to clean profile URL
            match = re.search(r"(https?://(?:www\.)?linkedin\.com/company/[^/?&#]+)", url)
            if match:
                return match.group(1)
    return ""


COMMON_EMAIL_PREFIXES = ["info", "contact", "admin", "enquiries", "enquire",
                         "hello", "mail", "office", "sales", "support"]


def find_email(company_name: str, domain: str = "") -> str:
    """
    Multi-strategy email finder. Fast: at most 2 web requests total.
      1. Snippet scan of one DuckDuckGo query (no page fetches)
      2. Fetch one directory/listing page if snippets found nothing
    """
    # ── Strategy 1: Scan DuckDuckGo snippets directly (single query) ──────────
    query = f'"{company_name}" email' if not domain else f'"{company_name}" "@{domain}"'
    emails = search_emails_from_snippets(query)

    if domain:
        hits = [e for e in emails if domain in e.lower()]
        if hits:
            return hits[0]
    elif emails:
        return emails[0]

    # ── Strategy 2: Fetch one non-skiplist page from a broader search ──────────
    fallback_query = f'"{company_name}" contact email'
    urls = search_web(fallback_query, num_results=4)
    for url in urls[:2]:
        parsed_domain = urlparse(url).netloc.lower().replace("www.", "")
        if any(skip in parsed_domain for skip in SKIP_DOMAINS):
            continue
        r = _get(url)
        if not r:
            continue
        page_emails = [e for e in EMAIL_RE.findall(r.text) if _is_valid_email(e)]
        if domain:
            hits = [e for e in page_emails if domain in e.lower()]
            if hits:
                return hits[0]
        if page_emails:
            return page_emails[0]
        break  # only try one page

    return ""


# ── Main enrichment function ───────────────────────────────────────────────────

def enrich_company(company_name: str, town: str = "", postcode: str = "") -> dict:
    """Return a dict of contact info for a single company name."""
    name = company_name.strip()
    location = " ".join(filter(None, [town, postcode, "UK"]))
    log.info(f"Enriching: {name}")

    result = {
        "company_name": name,
        "website":      "",
        "email":        "",
        "phone":        "",
        "linkedin":     "",
        "status":       "ok",
    }

    try:
        # 1. Google Places — use location for precision; treat as a hint, not final answer
        google_info = lookup_google_places(name, location=location)
        places_phone = google_info["phone"]   # held separately; website can override
        if google_info["website"] and verify_website_ownership(google_info["website"], name):
            result["website"] = google_info["website"]
            log.debug(f"Google Places found verified website for {name}")
        _sleep()

        # 2. Find website via search if Places didn't return one
        if not result["website"]:
            result["website"] = find_website(name, location=f"{town} {postcode}".strip())
            _sleep()

        website = result["website"]

        # 3. Scrape website — authoritative source; always overrides Places phone
        if website:
            contacts = scrape_website(website)
            if contacts["email"]:
                result["email"] = contacts["email"]
            if contacts["phone"]:
                result["phone"] = contacts["phone"]
                log.debug(f"Website phone used for {name}")
            elif places_phone:
                result["phone"] = places_phone
                log.debug(f"Google Places phone used (no website phone) for {name}")

            # Browser fallback only when HTML scraping found nothing at all
            if not result["email"] and not result["phone"]:
                log.debug(f"HTML scraping found nothing, trying browser for {name}")
                browser_contacts = scrape_website_browser(website)
                if browser_contacts["email"]:
                    result["email"] = browser_contacts["email"]
                if browser_contacts["phone"]:
                    result["phone"] = browser_contacts["phone"]

        else:
            # No website found — mine Places and Facebook directly for contacts
            log.debug(f"No website for {name}, trying Places + Facebook")
            if places_phone:
                result["phone"] = places_phone
            else:
                relaxed = lookup_google_places(name, location=location, threshold=0.3)
                if relaxed["phone"]:
                    result["phone"] = relaxed["phone"]
                    log.debug(f"Relaxed Places phone used for {name}")
            _sleep()

            fb_info = scrape_facebook_business(name)
            if fb_info.get("phone") and not result["phone"]:
                result["phone"] = fb_info["phone"]
            if fb_info.get("email"):
                result["email"] = fb_info["email"]

        # 4. Email still missing — multi-strategy web search
        if not result["email"]:
            log.debug(f"Running email search for {name}")
            domain = urlparse(result["website"]).netloc.replace("www.", "") if result["website"] else ""
            found_email = find_email(name, domain=domain)
            if found_email:
                result["email"] = found_email
                log.debug(f"Email found via web search for {name}: {found_email}")

        # 5. LinkedIn
        result["linkedin"] = find_linkedin(name)
        _sleep()

    except Exception as e:
        result["status"] = f"error: {e}"
        log.warning(f"Failed for {name}: {e}")

    return result


COMPANY_TIMEOUT = 120  # seconds per company before giving up


def enrich_company_safe(company_name: str, town: str = "", postcode: str = "") -> dict:
    """Run enrich_company with a hard 2-minute timeout. Logs and skips if exceeded."""
    from concurrent.futures import ThreadPoolExecutor, TimeoutError as FutureTimeout
    import time
    start = time.time()
    with ThreadPoolExecutor(max_workers=1) as ex:
        fut = ex.submit(enrich_company, company_name, town, postcode)
        try:
            result = fut.result(timeout=COMPANY_TIMEOUT)
            elapsed = time.time() - start
            if elapsed > 60:
                log.warning(f"SLOW ({elapsed:.0f}s): {company_name}")
            return result
        except FutureTimeout:
            elapsed = time.time() - start
            log.warning(f"TIMEOUT ({elapsed:.0f}s): {company_name} — skipped")
            return {
                "company_name": company_name,
                "website": "", "email": "", "phone": "", "linkedin": "",
                "status": f"timeout>{COMPANY_TIMEOUT}s",
            }


# ── Spreadsheet I/O ────────────────────────────────────────────────────────────

def google_sheets_to_csv_url(url: str) -> str:
    """Convert a Google Sheets URL to a CSV export URL."""
    # Extract sheet ID from URL like: https://docs.google.com/spreadsheets/d/SHEET_ID/edit?gid=GID
    import re
    sheet_match = re.search(r'/spreadsheets/d/([a-zA-Z0-9-_]+)', url)
    if not sheet_match:
        raise ValueError("Invalid Google Sheets URL")
    
    sheet_id = sheet_match.group(1)
    
    # Extract gid if present (sheet tab ID)
    gid_match = re.search(r'[?&#]gid=([0-9]+)', url)
    gid = gid_match.group(1) if gid_match else "0"
    
    return f"https://docs.google.com/spreadsheets/d/{sheet_id}/export?format=csv&gid={gid}"


def read_companies(filepath: str, column: str | None) -> list[dict]:
    """Read company names (plus Town/Postcode if present) from an xlsx, csv, or Google Sheets URL."""
    if "docs.google.com/spreadsheets" in filepath:
        log.info("Detected Google Sheets URL — converting to CSV export")
        filepath = google_sheets_to_csv_url(filepath)
        df = pd.read_csv(filepath)
    else:
        path = Path(filepath)
        if path.suffix.lower() == ".csv":
            df = pd.read_csv(filepath)
        else:
            df = pd.read_excel(filepath)

    if column:
        if column not in df.columns:
            raise ValueError(f"Column '{column}' not found. Available: {list(df.columns)}")
        name_col = column
    else:
        name_col = None
        for col in df.columns:
            sample = df[col].dropna().head(5).astype(str).tolist()
            if all(len(s) > 1 for s in sample):
                log.info(f"Using column: '{col}'")
                name_col = col
                break
        if name_col is None:
            name_col = df.columns[0]

    # Detect optional location columns
    town_col     = next((c for c in df.columns if c.strip().lower() == "town"), None)
    postcode_col = next((c for c in df.columns if "postcode" in c.strip().lower()), None)

    rows = []
    for _, row in df.iterrows():
        name = str(row[name_col]).strip()
        if not name or name == "nan":
            continue
        rows.append({
            "company_name": name,
            "town":     str(row[town_col]).strip() if town_col and pd.notna(row[town_col]) else "",
            "postcode": str(row[postcode_col]).strip() if postcode_col and pd.notna(row[postcode_col]) else "",
        })
    return rows


def write_results(results: list[dict], output_path: str):
    """Write enriched results to a temp file then atomically replace the target.
    This prevents crashes when the output file is open in Excel/Explorer."""
    import os, shutil
    tmp_path = output_path + ".tmp"

    df = pd.DataFrame(results)
    df.to_excel(tmp_path, index=False, engine="openpyxl")

    wb = load_workbook(tmp_path)
    ws = wb.active

    # Header styling
    header_fill = PatternFill("solid", start_color="1E3A5F")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    for cell in ws[1]:
        cell.fill      = header_fill
        cell.font      = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")

    # Column widths
    col_widths = {"A": 30, "B": 40, "C": 35, "D": 20, "E": 45, "F": 15}
    for col, width in col_widths.items():
        ws.column_dimensions[col].width = width
    ws.row_dimensions[1].height = 22

    # Row fills: red for no contacts found, zebra stripe otherwise
    light_fill = PatternFill("solid", start_color="EEF2F7")
    red_fill   = PatternFill("solid", start_color="FFCCCC")
    for row_idx, row in enumerate(ws.iter_rows(min_row=2), start=2):
        row_data   = {ws.cell(1, cell.column).value: cell.value for cell in row}
        no_contact = not any(row_data.get(f) for f in ("website", "email", "phone", "linkedin"))
        for cell in row:
            if no_contact:
                cell.fill = red_fill
            elif row_idx % 2 == 0:
                cell.fill = light_fill
            if cell.column == 2 and cell.value and str(cell.value).startswith("http"):
                cell.hyperlink = cell.value
                cell.font = Font(color="0563C1", underline="single")
            if cell.column == 6 and cell.value and cell.value != "ok":
                cell.font = Font(color="CC0000")

    ws.freeze_panes = "A2"
    wb.save(tmp_path)

    # Atomic replace: move temp over the real file
    try:
        if os.path.exists(output_path):
            os.replace(tmp_path, output_path)
        else:
            shutil.move(tmp_path, output_path)
        log.info(f"Results saved -> {output_path}")
    except PermissionError:
        log.warning(f"Output file locked — checkpoint kept as {tmp_path}")


# ── CLI ────────────────────────────────────────────────────────────────────────

def main():
    parser = argparse.ArgumentParser(
        description="Enrich company names with contact information."
    )
    parser.add_argument("input", help="Path to input .xlsx/.csv file or Google Sheets URL")
    parser.add_argument(
        "--column", "-c",
        default=None,
        help="Column name containing company names (auto-detected if omitted)",
    )
    parser.add_argument(
        "--output", "-o",
        default=None,
        help="Output file path (default: <input>_enriched.xlsx)",
    )
    parser.add_argument(
        "--start",
        type=int,
        default=1,
        help="Row number to start from (1-based, default: 1)",
    )
    parser.add_argument(
        "--limit", "-n",
        type=int,
        default=None,
        help="Number of companies to process (used with --start for a range)",
    )
    parser.add_argument(
        "--resume",
        default=None,
        help="Path to a partially-completed output file — skip already-enriched rows",
    )
    parser.add_argument(
        "--debug",
        action="store_true",
        help="Show detailed debug output (page URLs tried, what was found at each step)",
    )
    parser.add_argument(
        "--test",
        metavar="COMPANY",
        default=None,
        help="Test enrichment for a single company name and print results (no file needed)",
    )
    args = parser.parse_args()

    if args.debug:
        logging.getLogger().setLevel(logging.DEBUG)
        log.setLevel(logging.DEBUG)

    # ── Quick single-company test mode ────────────────────────────────────────
    if args.test:
        print(f"\nTesting enrichment for: {args.test!r}\n{'─'*50}")
        result = enrich_company(args.test)
        for k, v in result.items():
            print(f"  {k:15s}: {v or '(not found)'}")
        print(f"{'─'*50}\n")
        return

    # Resolve output path
    if args.output:
        output_path = args.output
    elif "docs.google.com/spreadsheets" in args.input:
        # For Google Sheets URLs, use a default name in current directory
        output_path = "enriched_companies.xlsx"
    else:
        input_path = Path(args.input)
        output_path = str(input_path.parent / (input_path.stem + "_enriched.xlsx"))

    # Read companies
    log.info(f"Reading companies from: {args.input}")
    companies = read_companies(args.input, args.column)
    log.info(f"Found {len(companies)} companies")

    start_idx = max(0, args.start - 1)  # convert 1-based to 0-based
    if start_idx:
        companies = companies[start_idx:]
        log.info(f"Starting from row {args.start}")
    if args.limit:
        companies = companies[: args.limit]
        log.info(f"Processing {args.limit} companies")

    # Resume support: skip already-done rows
    done: set[str] = set()
    results: list[dict] = []
    if args.resume and Path(args.resume).exists():
        existing = pd.read_excel(args.resume)
        done     = set(existing["company_name"].dropna().astype(str).tolist())
        results  = existing.to_dict("records")
        log.info(f"Resuming — {len(done)} already done, {len(companies)-len(done)} remaining")

    # Enrich (with parallel workers)
    to_process = [c for c in companies if c["company_name"] not in done]

    # Track original index so we can restore input order after parallel execution
    indexed = {i: c for i, c in enumerate(to_process)}
    new_results: dict[int, dict] = {}

    with ThreadPoolExecutor(max_workers=MAX_WORKERS) as executor:
        futures = {
            executor.submit(enrich_company_safe, c["company_name"], c.get("town", ""), c.get("postcode", "")): i
            for i, c in indexed.items()
        }

        with tqdm(total=len(to_process), desc="Enriching", unit="co") as pbar:
            for future in as_completed(futures):
                idx = futures[future]
                row = future.result()
                new_results[idx] = row
                pbar.update(1)

                # Write checkpoint every 25 rows (unordered is fine for checkpoints)
                if len(new_results) % 25 == 0:
                    checkpoint = results + [new_results[i] for i in sorted(new_results)]
                    write_results(checkpoint, output_path)
                    log.info(f"Checkpoint saved ({len(checkpoint)} rows)")

    # Merge: resumed rows first, then new rows in original input order
    results = results + [new_results[i] for i in sorted(new_results)]

    # Final save
    write_results(results, output_path)
    ok    = sum(1 for r in results if r["status"] == "ok")
    found = sum(1 for r in results if r.get("email") or r.get("phone"))
    print("\n" + "-" * 50)
    print(f"  Done!  {len(results)} companies processed")
    print(f"  {found} with at least one contact found")
    print(f"  {len(results)-ok} errors")
    print(f"  Output -> {output_path}")
    print("" + "-" * 50 + "\n")


if __name__ == "__main__":
    main()